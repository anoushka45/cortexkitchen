import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Brand palette ─────────────────────────────────────────────────────────────
_BG_HEADER   = "0A0F1C"   # near-black
_BG_SECTION  = "1E293B"   # slate-800
_BG_ROW_ALT  = "F0F4F8"   # light row alternation (light theme cells)
_FG_WHITE    = "FFFFFF"
_FG_AMBER    = "D97706"
_FG_SLATE    = "64748B"
_FG_DARK     = "1E293B"
_FG_GREEN    = "059669"
_FG_RED      = "DC2626"
_FG_YELLOW   = "D97706"

_VERDICT_COLOUR = {
    "approved": _FG_GREEN,
    "revision": _FG_YELLOW,
    "rejected": _FG_RED,
}

_PRIORITY_COLOUR = {
    "high":   _FG_RED,
    "medium": _FG_YELLOW,
    "low":    _FG_GREEN,
}

_THIN = Side(style="thin", color="E2E8F0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


# ── Public API ────────────────────────────────────────────────────────────────

def generate_run_excel(run: dict) -> bytes:
    """
    Build a role-aware .xlsx workbook for a planning run.

    Sheets
    ------
    1. Summary        — run metadata and critic verdict (all roles)
    2. Inventory & Staffing — shortage alerts and staffing actions (chef)
    3. Cost Breakdown  — LLM costs, token usage, dimension scores (owner)

    Parameters
    ----------
    run : dict
        Full planning run detail dict as returned by RunService.to_detail().

    Returns
    -------
    bytes
        Raw .xlsx bytes ready to stream as application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.
    """
    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    _build_summary_sheet(wb, run)
    _build_chef_sheet(wb, run)
    _build_owner_sheet(wb, run)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Sheet builders ────────────────────────────────────────────────────────────

def _build_summary_sheet(wb: Workbook, run: dict) -> None:
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False

    critic   = run.get("critic") or {}
    meta     = run.get("metadata") or {}
    verdict  = (critic.get("verdict") or "unknown").lower()
    score    = float(critic.get("score") or 0.0)
    scenario = (run.get("scenario") or "").replace("_", " ").title()

    # ── Title block ────────────────────────────────────────────────────────
    ws.merge_cells("A1:F1")
    _cell(ws, "A1", "CortexKitchen — Planning Run Summary",
          bold=True, size=16, fg=_FG_WHITE, bg=_BG_HEADER)

    ws.merge_cells("A2:F2")
    _cell(ws, "A2", f"Run #{run.get('id', '?')} · {scenario}",
          size=10, fg=_FG_SLATE, bg=_BG_HEADER)

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18

    # ── Meta table ─────────────────────────────────────────────────────────
    meta_rows = [
        ("Scenario",       scenario),
        ("Target Date",    run.get("target_date") or "—"),
        ("Generated At",   _fmt_dt(run.get("generated_at") or run.get("created_at"))),
        ("Status",         (run.get("status") or "—").title()),
        ("Critic Verdict", verdict.upper()),
        ("Critic Score",   f"{score:.2f} / 1.00"),
    ]
    start = 4
    _section_header(ws, f"A{start}", "Run Metadata", span="A:F")
    for i, (label, value) in enumerate(meta_rows, start + 1):
        _cell(ws, f"A{i}", label, bold=True, fg=_FG_SLATE, border=True)
        _cell(ws, f"B{i}", value,
              fg=_VERDICT_COLOUR.get(verdict, _FG_DARK) if label == "Critic Verdict" else _FG_DARK,
              bold=(label == "Critic Verdict"),
              border=True)

    # ── Dimension scores ───────────────────────────────────────────────────
    dims = critic.get("dimension_scores") or {}
    dim_keys = ["safety", "feasibility", "evidence", "actionability", "clarity"]
    row = start + len(meta_rows) + 2
    _section_header(ws, f"A{row}", "Critic Dimension Scores", span="A:F")
    row += 1
    for col_i, key in enumerate(dim_keys, 1):
        col = get_column_letter(col_i)
        _cell(ws, f"{col}{row}",     key.title(), bold=True, size=8, fg=_FG_SLATE, border=True)
        _cell(ws, f"{col}{row + 1}", f"{float(dims.get(key, 0)):.2f}", bold=True, fg=_FG_DARK, border=True)

    # ── Critic notes ───────────────────────────────────────────────────────
    row += 3
    notes = critic.get("notes") or ""
    if notes:
        _section_header(ws, f"A{row}", "Critic Notes", span="A:F")
        row += 1
        ws.merge_cells(f"A{row}:F{row + 2}")
        _cell(ws, f"A{row}", notes, fg=_FG_DARK, wrap=True)
        ws.row_dimensions[row].height = 48
        row += 3

    # ── Action items ───────────────────────────────────────────────────────
    feedback = critic.get("actionable_feedback") or []
    if feedback:
        _section_header(ws, f"A{row}", "Actionable Feedback", span="A:F")
        row += 1
        for item in feedback:
            ws.merge_cells(f"A{row}:F{row}")
            _cell(ws, f"A{row}", f"• {item}", fg=_FG_DARK, border=True, wrap=True)
            ws.row_dimensions[row].height = 30
            row += 1

    # ── LLM footer ─────────────────────────────────────────────────────────
    row += 1
    _section_header(ws, f"A{row}", "LLM Metadata", span="A:F")
    row += 1
    llm_rows = [
        ("Provider",   meta.get("llm_provider") or "—"),
        ("Model",      meta.get("llm_model") or "—"),
        ("Total Tokens", str(int(meta["total_tokens"])) if meta.get("total_tokens") is not None else "—"),
        ("Total Cost",   f"${float(meta['total_cost_usd']):.5f}" if meta.get("total_cost_usd") is not None else "—"),
        ("Duration",     f"{float(meta['total_duration_ms']) / 1000:.1f}s" if meta.get("total_duration_ms") is not None else "—"),
    ]
    for label, value in llm_rows:
        _cell(ws, f"A{row}", label, bold=True, fg=_FG_SLATE, border=True)
        _cell(ws, f"B{row}", value, fg=_FG_DARK, border=True)
        row += 1

    _set_column_widths(ws, [22, 28, 18, 18, 18, 18])


def _build_chef_sheet(wb: Workbook, run: dict) -> None:
    ws = wb.create_sheet("Inventory & Staffing")
    ws.sheet_view.showGridLines = False

    recs      = run.get("recommendations") or {}
    inventory = recs.get("inventory") or {}
    inv_data  = inventory.get("data") or {}
    forecast  = recs.get("forecast") or {}
    reserv    = recs.get("reservation") or {}

    ws.merge_cells("A1:G1")
    _cell(ws, "A1", "Inventory & Staffing — Chef View",
          bold=True, size=14, fg=_FG_WHITE, bg=_BG_HEADER)
    ws.row_dimensions[1].height = 26

    row = 3

    # ── Shortage alerts ────────────────────────────────────────────────────
    shortage_alerts = inv_data.get("shortage_alerts") or []
    _section_header(ws, f"A{row}", f"Shortage Alerts ({len(shortage_alerts)})", span="A:G")
    row += 1
    headers = ["Ingredient", "In Stock", "Unit", "Threshold", "Shortfall", "Severity", "Spoilage Risk"]
    for col_i, h in enumerate(headers, 1):
        _cell(ws, f"{get_column_letter(col_i)}{row}", h,
              bold=True, size=9, fg=_FG_WHITE, bg=_BG_SECTION, border=True)
    row += 1
    if shortage_alerts:
        for alert in shortage_alerts:
            severity    = (alert.get("severity") or "").lower()
            row_fg      = _PRIORITY_COLOUR.get(severity, _FG_DARK)
            sev_display = severity.upper() if severity else "—"
            values = [
                alert.get("ingredient") or "—",
                alert.get("quantity_in_stock"),
                alert.get("unit") or "—",
                alert.get("reorder_threshold"),
                alert.get("shortfall"),
                sev_display,
                "Yes" if alert.get("spoilage_risk") else "No",
            ]
            for col_i, v in enumerate(values, 1):
                col_letter = get_column_letter(col_i)
                is_sev = col_i == 6
                _cell(ws, f"{col_letter}{row}", v if v is not None else "—",
                      fg=row_fg if is_sev else _FG_DARK,
                      bold=is_sev, border=True)
            row += 1
    else:
        ws.merge_cells(f"A{row}:G{row}")
        _cell(ws, f"A{row}", "No shortage alerts.", fg=_FG_SLATE, border=True)
        row += 1

    # ── Overstock alerts ───────────────────────────────────────────────────
    overstock_alerts = inv_data.get("overstock_alerts") or []
    row += 1
    _section_header(ws, f"A{row}", f"Overstock Alerts ({len(overstock_alerts)})", span="A:G")
    row += 1
    if overstock_alerts:
        ov_headers = ["Ingredient", "In Stock", "Unit", "Overstock By", "Waste Risk"]
        for col_i, h in enumerate(ov_headers, 1):
            _cell(ws, f"{get_column_letter(col_i)}{row}", h,
                  bold=True, size=9, fg=_FG_WHITE, bg=_BG_SECTION, border=True)
        row += 1
        for alert in overstock_alerts:
            values = [
                alert.get("ingredient") or "—",
                alert.get("quantity_in_stock"),
                alert.get("unit") or "—",
                alert.get("overstock_amount"),
                "Yes" if alert.get("waste_risk") else "No",
            ]
            for col_i, v in enumerate(values, 1):
                _cell(ws, f"{get_column_letter(col_i)}{row}", v if v is not None else "—",
                      fg=_FG_DARK, border=True)
            row += 1
    else:
        ws.merge_cells(f"A{row}:G{row}")
        _cell(ws, f"A{row}", "No overstock alerts.", fg=_FG_SLATE, border=True)
        row += 1

    # ── Restock actions ────────────────────────────────────────────────────
    restock = inventory.get("restock_actions") or []
    row += 1
    _section_header(ws, f"A{row}", "Restock Actions", span="A:G")
    row += 1
    if restock:
        for action in restock:
            ws.merge_cells(f"A{row}:G{row}")
            _cell(ws, f"A{row}", f"• {action}", fg=_FG_DARK, border=True, wrap=True)
            ws.row_dimensions[row].height = 22
            row += 1
    else:
        ws.merge_cells(f"A{row}:G{row}")
        _cell(ws, f"A{row}", "No restock actions.", fg=_FG_SLATE, border=True)
        row += 1

    # ── Staffing & covers ──────────────────────────────────────────────────
    row += 1
    _section_header(ws, f"A{row}", "Staffing & Covers", span="A:G")
    row += 1
    staffing_rows = []
    if forecast.get("predicted_covers"):
        staffing_rows.append(("Predicted Covers", str(forecast["predicted_covers"])))
    if forecast.get("peak_hours"):
        staffing_rows.append(("Peak Hours", ", ".join(forecast["peak_hours"]) if isinstance(forecast["peak_hours"], list) else str(forecast["peak_hours"])))
    if reserv.get("covers"):
        staffing_rows.append(("Reserved Covers", str(reserv["covers"])))
    if reserv.get("peak_hour"):
        staffing_rows.append(("Reservation Peak Hour", str(reserv["peak_hour"])))
    if reserv.get("waitlist_risk"):
        staffing_rows.append(("Waitlist Risk", str(reserv["waitlist_risk"]).title()))
    if reserv.get("recommendation"):
        staffing_rows.append(("Staffing Recommendation", str(reserv["recommendation"])[:300]))

    if staffing_rows:
        for label, value in staffing_rows:
            _cell(ws, f"A{row}", label, bold=True, fg=_FG_SLATE, border=True)
            ws.merge_cells(f"B{row}:G{row}")
            _cell(ws, f"B{row}", value, fg=_FG_DARK, border=True, wrap=True)
            ws.row_dimensions[row].height = 22
            row += 1
    else:
        ws.merge_cells(f"A{row}:G{row}")
        _cell(ws, f"A{row}", "No staffing data available.", fg=_FG_SLATE, border=True)

    _set_column_widths(ws, [22, 14, 10, 14, 14, 14, 12])


def _build_owner_sheet(wb: Workbook, run: dict) -> None:
    ws = wb.create_sheet("Cost Breakdown")
    ws.sheet_view.showGridLines = False

    critic       = run.get("critic") or {}
    meta         = run.get("metadata") or {}
    cost_analysis = critic.get("cost_analysis") or {}

    ws.merge_cells("A1:E1")
    _cell(ws, "A1", "Cost Breakdown — Owner View",
          bold=True, size=14, fg=_FG_WHITE, bg=_BG_HEADER)
    ws.row_dimensions[1].height = 26

    row = 3

    # ── LLM usage ──────────────────────────────────────────────────────────
    _section_header(ws, f"A{row}", "LLM Usage", span="A:E")
    row += 1
    usage_rows = [
        ("Provider",       meta.get("llm_provider") or "—"),
        ("Model",          meta.get("llm_model") or "—"),
        ("Prompt Tokens",  str(int(meta["prompt_tokens"])) if meta.get("prompt_tokens") is not None else "—"),
        ("Completion Tokens", str(int(meta["completion_tokens"])) if meta.get("completion_tokens") is not None else "—"),
        ("Total Tokens",   str(int(meta["total_tokens"])) if meta.get("total_tokens") is not None else "—"),
        ("Total Cost (USD)", f"${float(meta['total_cost_usd']):.5f}" if meta.get("total_cost_usd") is not None else "—"),
        ("Duration",       f"{float(meta['total_duration_ms']) / 1000:.1f}s" if meta.get("total_duration_ms") is not None else "—"),
    ]
    for label, value in usage_rows:
        _cell(ws, f"A{row}", label, bold=True, fg=_FG_SLATE, border=True)
        _cell(ws, f"B{row}", value, fg=_FG_DARK, border=True)
        row += 1

    # ── Critic dimension scores ────────────────────────────────────────────
    dims     = critic.get("dimension_scores") or {}
    dim_keys = ["safety", "feasibility", "evidence", "actionability", "clarity"]
    row += 1
    _section_header(ws, f"A{row}", "Critic Dimension Scores", span="A:E")
    row += 1
    _cell(ws, f"A{row}", "Dimension",  bold=True, size=9, fg=_FG_WHITE, bg=_BG_SECTION, border=True)
    _cell(ws, f"B{row}", "Score",      bold=True, size=9, fg=_FG_WHITE, bg=_BG_SECTION, border=True)
    _cell(ws, f"C{row}", "Score /100", bold=True, size=9, fg=_FG_WHITE, bg=_BG_SECTION, border=True)
    row += 1
    for key in dim_keys:
        raw   = float(dims.get(key, 0))
        pct   = round(raw * 100)
        color = _FG_GREEN if pct >= 75 else (_FG_YELLOW if pct >= 50 else _FG_RED)
        _cell(ws, f"A{row}", key.title(), fg=_FG_DARK, border=True)
        _cell(ws, f"B{row}", f"{raw:.2f}", fg=_FG_DARK, border=True)
        _cell(ws, f"C{row}", f"{pct}/100", fg=color, bold=True, border=True)
        row += 1

    # ── Cost-aware analysis ────────────────────────────────────────────────
    if cost_analysis:
        row += 1
        _section_header(ws, f"A{row}", "Cost-Aware Analysis", span="A:E")
        row += 1
        ca_rows = [
            ("Cost Pressure Score", f"{float(cost_analysis.get('cost_pressure_score', 0)):.2f}"),
            ("Benefit Score",       f"{float(cost_analysis.get('benefit_score', 0)):.2f}"),
            ("Tradeoff Score",      f"{float(cost_analysis.get('tradeoff_score', 0)):.2f}"),
        ]
        for label, value in ca_rows:
            _cell(ws, f"A{row}", label, bold=True, fg=_FG_SLATE, border=True)
            _cell(ws, f"B{row}", value, fg=_FG_DARK, border=True)
            row += 1

    # ── Overall verdict ────────────────────────────────────────────────────
    verdict = (critic.get("verdict") or "unknown").lower()
    score   = float(critic.get("score") or 0.0)
    row += 1
    _section_header(ws, f"A{row}", "Overall Verdict", span="A:E")
    row += 1
    _cell(ws, f"A{row}", "Verdict", bold=True, fg=_FG_SLATE, border=True)
    _cell(ws, f"B{row}", verdict.upper(),
          fg=_VERDICT_COLOUR.get(verdict, _FG_DARK), bold=True, border=True)
    row += 1
    _cell(ws, f"A{row}", "Score", bold=True, fg=_FG_SLATE, border=True)
    _cell(ws, f"B{row}", f"{score:.2f} / 1.00", bold=True, fg=_FG_DARK, border=True)

    _set_column_widths(ws, [26, 18, 14, 14, 14])


# ── Style helpers ─────────────────────────────────────────────────────────────

def _cell(ws, coord: str, value, *, bold=False, size=10, fg=_FG_DARK,
          bg: str | None = None, border=False, wrap=False) -> None:
    c = ws[coord]
    c.value = value
    c.font  = Font(name="Calibri", bold=bold, size=size,
                   color=fg if not bg else _FG_WHITE)
    if bg:
        c.fill = PatternFill(fill_type="solid", fgColor=bg)
    if border:
        c.border = _BORDER
    c.alignment = Alignment(
        vertical="center",
        horizontal="left",
        wrap_text=wrap,
    )


def _section_header(ws, coord: str, label: str, *, span: str = "") -> None:
    c = ws[coord]
    c.value = label
    c.font  = Font(name="Calibri", bold=True, size=9,
                   color=_FG_WHITE)
    c.fill  = PatternFill(fill_type="solid", fgColor=_BG_SECTION)
    c.alignment = Alignment(vertical="center", horizontal="left")
    if span:
        col_start = coord[0]
        row_num   = coord[1:]
        col_end   = span.split(":")[1]
        ws.merge_cells(f"{col_start}{row_num}:{col_end}{row_num}")
    ws.row_dimensions[int(coord[1:])].height = 20


def _set_column_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _fmt_dt(value: str | None) -> str:
    if not value:
        return "—"
    try:
        dt = datetime.fromisoformat(value.replace("Z", "+00:00"))
        return dt.strftime("%d %b %Y, %H:%M")
    except Exception:
        return str(value)[:16]
