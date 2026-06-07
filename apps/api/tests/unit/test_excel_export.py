"""Unit tests for P5-02 Excel export — report_generator."""

import io
import pytest
from openpyxl import load_workbook
from app.infrastructure.excel.report_generator import generate_run_excel

MOCK_RUN = {
    "id": 42,
    "scenario": "friday_rush",
    "target_date": "2026-06-06",
    "status": "ready",
    "generated_at": "2026-06-03T10:00:00+00:00",
    "created_at": "2026-06-03T10:01:00+00:00",
    "critic_verdict": "approved",
    "critic_score": 0.92,
    "critic": {
        "verdict": "approved",
        "score": 0.92,
        "notes": "The plan addresses high occupancy and critical inventory shortages.",
        "dimension_scores": {
            "safety": 1.0,
            "feasibility": 0.95,
            "evidence": 0.90,
            "actionability": 0.90,
            "clarity": 0.90,
        },
        "actionable_feedback": [
            "Prioritise critical shortage restocking before service opens.",
            "Monitor staffing levels during peak hours.",
        ],
        "revision_reasons": [],
        "cost_analysis": {
            "cost_pressure_score": 0.72,
            "benefit_score": 0.88,
            "tradeoff_score": 0.80,
        },
    },
    "recommendations": {
        "forecast": {
            "recommendation": "Increase staffing by 15% and prep 25% more for top items.",
            "predicted_covers": 180,
            "peak_hours": ["18:00", "19:30"],
            "priority": "high",
        },
        "reservation": {
            "recommendation": "Monitor and adjust seating assignments as needed.",
            "covers": 160,
            "peak_hour": "19:00",
            "waitlist_risk": "high",
            "priority": "medium",
        },
        "complaint": {
            "overall_summary": "Main issues are stockouts and slow prep.",
            "priority": "high",
        },
        "menu": {
            "reasoning": "Focus on popular burger items.",
            "priority": "medium",
        },
        "inventory": {
            "restock_actions": [
                "Order 10kg Mozzarella immediately",
                "Reorder burger buns — 2 days supply remaining",
            ],
            "waste_reduction_actions": [],
            "priority": "high",
            "reasoning": "Critical shortage on high-demand week.",
            "risks": ["Unable to fulfil pizza orders during peak hours"],
            "data": {
                "shortage_alerts": [
                    {
                        "ingredient": "Mozzarella",
                        "unit": "kg",
                        "quantity_in_stock": 3.5,
                        "reorder_threshold": 8.0,
                        "shortfall": 4.5,
                        "spoilage_risk": True,
                        "severity": "critical",
                    }
                ],
                "overstock_alerts": [],
            },
        },
    },
    "metadata": {
        "run_id": "abc12345",
        "llm_provider": "groq",
        "llm_model": "llama-3.3-70b-versatile",
        "total_cost_usd": 0.0038,
        "total_tokens": 5800,
        "prompt_tokens": 4200,
        "completion_tokens": 1600,
        "total_duration_ms": 8500,
    },
    "rag_context": {},
    "final_response": {},
}


def _load(run: dict):
    """Generate workbook bytes and return an openpyxl Workbook."""
    return load_workbook(io.BytesIO(generate_run_excel(run)))


def _all_values(ws) -> set[str]:
    """Collect all non-empty string cell values in a sheet."""
    values = set()
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if v is not None:
                values.add(str(v))
    return values


# ── Basic structure ───────────────────────────────────────────────────────────

def test_returns_bytes():
    result = generate_run_excel(MOCK_RUN)
    assert isinstance(result, bytes)
    assert len(result) > 1000


def test_xlsx_magic_bytes():
    result = generate_run_excel(MOCK_RUN)
    # .xlsx files are ZIP archives starting with PK\x03\x04
    assert result[:2] == b"PK"


def test_has_three_sheets():
    wb = _load(MOCK_RUN)
    assert len(wb.sheetnames) == 3


def test_sheet_names():
    wb = _load(MOCK_RUN)
    assert wb.sheetnames == ["Summary", "Inventory & Staffing", "Cost Breakdown"]


# ── Summary sheet ─────────────────────────────────────────────────────────────

def test_summary_contains_scenario():
    ws = _load(MOCK_RUN)["Summary"]
    values = _all_values(ws)
    assert any("Friday Rush" in v for v in values)


def test_summary_contains_verdict():
    ws = _load(MOCK_RUN)["Summary"]
    values = _all_values(ws)
    assert "APPROVED" in values


def test_summary_contains_critic_score():
    ws = _load(MOCK_RUN)["Summary"]
    values = _all_values(ws)
    assert any("0.92" in v for v in values)


def test_summary_contains_all_dimensions():
    ws = _load(MOCK_RUN)["Summary"]
    values = _all_values(ws)
    for dim in ["Safety", "Feasibility", "Evidence", "Actionability", "Clarity"]:
        assert dim in values, f"Missing dimension: {dim}"


def test_summary_contains_llm_provider():
    ws = _load(MOCK_RUN)["Summary"]
    values = _all_values(ws)
    assert any("groq" in v.lower() for v in values)


def test_summary_contains_action_items():
    ws = _load(MOCK_RUN)["Summary"]
    values = _all_values(ws)
    assert any("restocking" in v.lower() for v in values)


# ── Inventory & Staffing sheet ────────────────────────────────────────────────

def test_chef_sheet_contains_shortage_alert():
    ws = _load(MOCK_RUN)["Inventory & Staffing"]
    values = _all_values(ws)
    assert any("Mozzarella" in v for v in values)


def test_chef_sheet_contains_severity():
    ws = _load(MOCK_RUN)["Inventory & Staffing"]
    values = _all_values(ws)
    assert "CRITICAL" in values


def test_chef_sheet_contains_restock_action():
    ws = _load(MOCK_RUN)["Inventory & Staffing"]
    values = _all_values(ws)
    assert any("Mozzarella" in v or "restock" in v.lower() for v in values)


def test_chef_sheet_contains_staffing_data():
    ws = _load(MOCK_RUN)["Inventory & Staffing"]
    values = _all_values(ws)
    # predicted covers or reserved covers should appear
    assert any("180" in v or "160" in v for v in values)


def test_chef_sheet_contains_peak_hours():
    ws = _load(MOCK_RUN)["Inventory & Staffing"]
    values = _all_values(ws)
    assert any("18:00" in v or "19:00" in v for v in values)


# ── Cost Breakdown sheet ──────────────────────────────────────────────────────

def test_owner_sheet_contains_cost():
    ws = _load(MOCK_RUN)["Cost Breakdown"]
    values = _all_values(ws)
    assert any("0.00380" in v for v in values)


def test_owner_sheet_contains_token_counts():
    ws = _load(MOCK_RUN)["Cost Breakdown"]
    values = _all_values(ws)
    assert any("5800" in v for v in values)


def test_owner_sheet_contains_cost_analysis():
    ws = _load(MOCK_RUN)["Cost Breakdown"]
    values = _all_values(ws)
    assert any("0.72" in v for v in values)  # cost_pressure_score


def test_owner_sheet_contains_verdict():
    ws = _load(MOCK_RUN)["Cost Breakdown"]
    values = _all_values(ws)
    assert "APPROVED" in values


def test_owner_sheet_contains_dimension_scores():
    ws = _load(MOCK_RUN)["Cost Breakdown"]
    values = _all_values(ws)
    for dim in ["Safety", "Feasibility", "Evidence", "Actionability", "Clarity"]:
        assert dim in values, f"Missing dimension: {dim}"


# ── Resilience ────────────────────────────────────────────────────────────────

def test_handles_missing_optional_fields():
    minimal = {
        "id": 1,
        "scenario": "weekday_lunch",
        "target_date": None,
        "status": "ready",
        "generated_at": None,
        "created_at": None,
        "critic": {
            "verdict": "revision",
            "score": 0.5,
            "notes": "",
            "dimension_scores": {},
            "actionable_feedback": [],
        },
        "recommendations": {},
        "metadata": {},
        "rag_context": {},
        "final_response": {},
    }
    wb = _load(minimal)
    assert len(wb.sheetnames) == 3


def test_handles_empty_inventory_data():
    run = {**MOCK_RUN, "recommendations": {**MOCK_RUN["recommendations"],
        "inventory": {"restock_actions": [], "data": {}}}}
    wb = _load(run)
    ws = wb["Inventory & Staffing"]
    values = _all_values(ws)
    assert any("No shortage alerts" in v for v in values)


def test_rejected_verdict_generates_without_error():
    run = {**MOCK_RUN, "critic": {**MOCK_RUN["critic"], "verdict": "rejected", "score": 0.1}}
    wb = _load(run)
    ws = wb["Cost Breakdown"]
    values = _all_values(ws)
    assert "REJECTED" in values
